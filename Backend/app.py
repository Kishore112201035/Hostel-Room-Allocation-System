import openpyxl
import importlib.util
import sys
import os

# --- File paths ---
excel_path = "hostel_data.xlsx"
rooms_file = "rooms.py"

# --- Check if files exist ---
if not os.path.exists(excel_path):
    raise FileNotFoundError(f"❌ Excel file not found: {excel_path}")
if not os.path.exists(rooms_file):
    raise FileNotFoundError(f"❌ Rooms file not found: {rooms_file}")

# --- Load Excel ---
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active   # first sheet (with student data)

# Remove old allocation sheets if exist
for sname in ["women allocation", "men allocation"]:
    if sname in wb.sheetnames:
        ws = wb[sname]
        wb.remove(ws)

# Create fresh allocation sheets
ws_women = wb.create_sheet("women allocation")
ws_men = wb.create_sheet("men allocation")

ws_women.append(["Category", "Email ID", "Allocated Room"])
ws_men.append(["Category", "Email ID", "Allocated Room"])

# --- Load rooms.py dynamically ---
spec = importlib.util.spec_from_file_location("rooms", rooms_file)
rooms = importlib.util.module_from_spec(spec)
sys.modules["rooms"] = rooms
spec.loader.exec_module(rooms)

# ----------------------------------------------------------
# ✅ Separate blocks for Women & Men
# ----------------------------------------------------------
WOMEN_BLOCKS = {
    "AM": "MF",   # Acad Malhar → MF
    "MF": "MF",   # direct code
    "AF": "AF",   # Acad-facing
    "FF": "FF",   # Forest-facing
    "HF": "HF",   # Hill-facing
    "SA": "AF"    # Saveri Acad → map to AF
}

MEN_BLOCKS = {
    "SMS": "SMS",
    "SOS": "SOS",
    "SKS": "SKS",
    "SAS": "SAS",
    "MF1S": "MF1S",
    "MF2S": "MF2S",
    "MBS": "MBS"
}

# ----------------------------------------------------------
# ✅ Allocation function (handles gender)
# ----------------------------------------------------------
def allocate_room(gender, pref_code, floor_num):
    """Allocate first free room based on gender, block preference and floor."""
    block_map = WOMEN_BLOCKS if gender.lower() == "women" else MEN_BLOCKS

    pref_code = str(pref_code).strip().upper()
    if pref_code not in block_map:
        return None

    block_name = block_map[pref_code]
    block = getattr(rooms, block_name, None)
    if not block:
        return None

    try:
        floor_idx = int(floor_num) - 1  # Convert floor number to index
    except Exception:
        return None

    if floor_idx < 0 or floor_idx >= len(block):
        return None

    # Find first unallocated room
    for room in block[floor_idx]:
        if room[1] == 0:
            room[1] = 1
            return room[0]
    return None

# ----------------------------------------------------------
# ✅ Allocation loop (writes into correct sheet)
# ----------------------------------------------------------
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
    if not row or not row[0].value:
        continue

    category = row[0].value   # ⚠️ "Category" instead of "Gender"
    allocated_cell = row[1]   # ⚠️ "Allocated" column (to be updated)
    email = row[2].value

    allocated = None

    # ------------------------------------------------------
    # ✅ Loop through all (FLOORx, PREFx) pairs until blank
    # Columns start at index 6 (FLOOR1) → then PREF1, etc.
    # ------------------------------------------------------
    col = 6
    while col + 1 < len(row):
        floor_val = row[col].value
        pref_val = row[col + 1].value

        if not pref_val or not floor_val:  # stop if blank
            break  

        if not allocated:  # only allocate if not already done
            allocated = allocate_room(category, pref_val, floor_val)

        col += 2  # move to next (FLOORx, PREFx)

    # Update "Allocated" column → 1 if allocated else 0
    allocated_cell.value = 1 if allocated else 0

    # Write into separate sheets
    if category.lower() == "women":
        ws_women.append([category, email, allocated if allocated else "Not Allocated"])
    else:
        ws_men.append([category, email, allocated if allocated else "Not Allocated"])

# Save updated Excel
wb.save(excel_path)

# ----------------------------------------------------------
# ✅ Overwrite rooms.py with updated allocations
# ----------------------------------------------------------
with open(rooms_file, "w") as f:
    f.write("#Women\n")
    f.write("# Savari Malhar Facing (AM)\n")
    f.write(f"MF = {rooms.MF}\n\n")

    f.write("# Saveri Acad-facing rooms (SA)\n")
    f.write(f"AF = {rooms.AF}\n\n")

    f.write("#Malhar Forest Facing (FF)\n")
    f.write(f"FF = {rooms.FF}\n\n")

    f.write("#Malhar Hill Facing (HF)\n")
    f.write(f"HF = {rooms.HF}\n\n")

    f.write("\n\n#MEN\n")
    f.write("#Savari Malhar Side (SMS)\n")
    f.write(f"SMS = {rooms.SMS}\n\n")

    f.write("#Saveri Opean Side (SOS)\n")
    f.write(f"SOS = {rooms.SOS}\n\n")

    f.write("#Saveri kadaram side (SKS)\n")
    f.write(f"SKS = {rooms.SKS}\n\n")

    f.write("#Saveri ACad Side (SAS)\n")
    f.write(f"SAS = {rooms.SAS}\n\n")

    f.write("#Mahlar Forest 1 side (MF1S)\n")
    f.write(f"MF1S = {rooms.MF1S}\n\n")

    f.write("#Malhar Forest 2 Side (MF2S)\n")
    f.write(f"MF2S = {rooms.MF2S}\n\n")

    f.write("#Malhar BasketBall Court Side\n")
    f.write(f"MBS = {rooms.MBS}\n\n")

    # ✅ Keep print_allocation() in rooms.py
    f.write("""
def print_allocation():
    print("\\n--- Room Allocation Status ---")
    print("Choose option:")
    print("1 → Show both Allocated & Unallocated")
    print("2 → Show only Allocated")
    print("3 → Show only Unallocated")

    choice = input("Enter 1, 2, or 3: ").strip()

    blocks = {
        "MF": MF,
        "AF": AF,
        "FF": FF,
        "HF": HF,
        "SMS": SMS,
        "SOS": SOS,
        "SKS": SKS,
        "SAS": SAS,
        "MF1S": MF1S,
        "MF2S": MF2S,
        "MBS": MBS,
    }

    for block_name, block in blocks.items():
        print(f"\\nBlock {block_name}:")
        for floor_idx, floor in enumerate(block):
            if floor_idx == 0:
                floor_name = "0th"
            elif floor_idx == 1:
                floor_name = "1st"
            elif floor_idx == 2:
                floor_name = "2nd"
            elif floor_idx == 3:
                floor_name = "3rd"
            else:
                floor_name = f"{floor_idx}th"

            allocated = [room[0] for room in floor if room[1] == 1]
            unallocated = [room[0] for room in floor if room[1] == 0]

            print(f"  Floor {floor_name} floor:")

            if choice == "1":  # both
                print(f"    Allocated   : {allocated if allocated else 'None'}")
                print(f"    Unallocated : {unallocated if unallocated else 'None'}")
            elif choice == "2":  # only allocated
                print(f"    Allocated   : {allocated if allocated else 'None'}")
            elif choice == "3":  # only unallocated
                print(f"    Unallocated : {unallocated if unallocated else 'None'}")
            else:
                print("    Invalid choice! Showing both by default.")
                print(f"    Allocated   : {allocated if allocated else 'None'}")
                print(f"    Unallocated : {unallocated if unallocated else 'None'}")

if __name__ == "__main__":
    print_allocation()
""")

print("✅ Allocation completed. Results saved in Excel + rooms.py updated.")
